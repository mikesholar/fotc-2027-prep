import json, re, pathlib, datetime, math
import openpyxl

ROOT = pathlib.Path(__file__).resolve().parents[1]
XLSX = ROOT / "FOTC_2027_26-Week_Prep_Mike-2.xlsx"
OUT = ROOT / "src" / "data" / "plan.json"

BLOCK_SHEETS = [
    "B1 Base W1-5", "B2 Build W6-10", "B3 Qualifier W11-15",
    "B4 Rebuild W16-20", "B5 Champ Prep W21-24", "B6 Peak W25-26",
]
BLOCK_LABEL = {
    "B1 Base W1-5": "B1 Base", "B2 Build W6-10": "B2 Build",
    "B3 Qualifier W11-15": "B3 Qualifier", "B4 Rebuild W16-20": "B4 Rebuild",
    "B5 Champ Prep W21-24": "B5 Champ Prep", "B6 Peak W25-26": "B6 Peak",
}

DEFAULT_MAXES = {
    "backSquat": 275, "frontSquat": 225, "bench": 195, "strictPress": 135,
    "deadlift": 315, "cleanJerk": 205, "snatch": 145,
}
LIFTS = [
    {"key": "backSquat", "name": "Back Squat", "isEstimate": False},
    {"key": "frontSquat", "name": "Front Squat", "isEstimate": False},
    {"key": "bench", "name": "Bench Press", "isEstimate": False},
    {"key": "strictPress", "name": "Strict Press", "isEstimate": False},
    {"key": "deadlift", "name": "Deadlift", "isEstimate": False},
    {"key": "cleanJerk", "name": "Clean & Jerk", "isEstimate": True},
    {"key": "snatch", "name": "Snatch", "isEstimate": True},
]

MONTHS = {m: i for i, m in enumerate(
    ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"], start=1)}

WEEK_HDR = re.compile(r"WEEK\s+(\d+)\s*·\s*([^—]+?)\s*—\s*(.*)", re.UNICODE)

# Barbell lift names as they lead a main-lift line, mapped to a lift key. Longer
# names come first so "BENCH PRESS" wins over "BENCH". PUSH PRESS shares the
# strictPress max; C&J shares cleanJerk.
LIFT_BY_NAME = [
    ("BACK SQUAT", "backSquat"), ("FRONT SQUAT", "frontSquat"),
    ("BENCH PRESS", "bench"), ("BENCH", "bench"),
    ("STRICT PRESS", "strictPress"), ("PUSH PRESS", "strictPress"),
    ("DEADLIFT", "deadlift"), ("CLEAN & JERK", "cleanJerk"), ("C&J", "cleanJerk"),
    ("SNATCH", "snatch"),
]
DISPLAY_NAME = {l["key"]: l["name"] for l in LIFTS}

SET_RE = re.compile(r"(\d+)\s*x\s*(\d+)\s*@\s*(\d+(?:\.\d+)?)%", re.IGNORECASE)
PCT_RE = re.compile(r"@\s*(\d+(?:\.\d+)?)%")

# Boundary markers: a "·"-segment starting with any of these (or a DIFFERENT lift
# name) is an accessory / different movement, not part of the main lift.
STOP_RE = re.compile(r"^(\+|E:30|E90s|EMOM|E2MOM|EOM|AMRAP|\d+:\d+)", re.IGNORECASE)

# A line's first "·"-segment counts as a strength prescription if it carries any of
# these signals: a rep scheme (NxM / N/N/N), an "@ P%", or a heavy/climb cue.
STRENGTH_RE = re.compile(
    r"\d+\s*[x/]\s*\d+|@\s*\d+(?:\.\d+)?%|\bheavy\b|\bclimb|\bmax\b|\bbuild\b|\bmoderate\b",
    re.IGNORECASE)

# Leading rep-scheme token for a note set: "5/5/3/3/3", "1x1", "3/3/1/1/1".
NOTE_SCHEME_RE = re.compile(r"^((?:\d+\s*x\s*\d+)|(?:[\d/]+))\s*[—-]?\s*(.*)$")


def round5(x):
    # Round-half-UP to nearest 5 to match the spreadsheet (Excel ROUND) and the
    # app's TS Math.round. Python's round() is banker's rounding and diverges at
    # .5 boundaries (e.g. 0.9*225=202.5 -> sheet prints 205, not 200).
    return int(math.floor(x / 5 + 0.5)) * 5


def parse_load_cell(load_cell):
    # "80% → 220 lb · 82.5% → 225 lb" -> [(0.80, 220), (0.825, 225), ...]
    pairs = []
    for pct_s, load_s in re.findall(r"(\d+(?:\.\d+)?)%\s*→\s*(\d+)", load_cell):
        pairs.append((float(pct_s) / 100.0, int(load_s)))
    return pairs


def backsolve_lift(load_cell):
    # The lift whose default max reproduces every printed (pct -> load) pair under
    # round-half-up. This tells us which lift the sheet's load cell was auto-computed
    # for -- but the cell may belong to an accessory, so it is only ONE signal.
    pairs = parse_load_cell(load_cell)
    if not pairs:
        return None
    for key, mx in DEFAULT_MAXES.items():
        if all(round5(p * mx) == load for p, load in pairs):
            return key
    return None


def find_lead_lift(session):
    # leadLift = the first non-prep line after the title that leads (optionally after
    # an "NxM " scheme) with a barbell lift name AND whose first "·"-segment is a real
    # strength prescription. Line-start anchoring keeps mid-line metcon mentions
    # (e.g. "20 DB snatch 50") from ever counting as a main lift.
    for line in session.splitlines()[1:]:
        s = line.strip()
        if s.lower().startswith("prep"):
            continue
        first_seg = s.split("·")[0]
        if not STRENGTH_RE.search(first_seg):
            continue
        for name, key in LIFT_BY_NAME:
            if re.match(r"(?:\d+\s*x\s*\d+\s+)?" + re.escape(name) + r"\b", s, re.IGNORECASE):
                return s, name, key
    return None, None, None


def find_pct_line(session, key):
    # The line that prescribes `key` with an "@ P%" -- used when the load cell resolves
    # to `key` (backsolve). Metcons use absolute loads, so requiring "@ P%" alongside
    # the lift name avoids picking up conditioning mentions of the lift.
    aliases = [name for name, k in LIFT_BY_NAME if k == key]
    for line in session.splitlines()[1:]:
        s = line.strip()
        if s.lower().startswith("prep"):
            continue
        if PCT_RE.search(s):
            for name in aliases:
                if re.search(re.escape(name) + r"\b", s, re.IGNORECASE):
                    return s, name
    return None, None


def parse_note_segment(seg):
    seg = seg.strip(" ·—-.")
    m = NOTE_SCHEME_RE.match(seg)
    if m and m.group(1):
        return {"scheme": re.sub(r"\s*x\s*", "×", m.group(1)), "note": m.group(2).strip(" —-·.")}
    return {"scheme": "", "note": seg}


def parse_piece(piece, mx):
    piece = piece.strip(" ·—.")
    if not piece:
        return []
    contiguous = SET_RE.findall(piece)
    if contiguous:
        return [{"scheme": f"{c}×{r}", "pct": float(p) / 100.0,
                 "expectedLoad": round5(float(p) / 100.0 * mx)} for c, r, p in contiguous]
    pcts = PCT_RE.findall(piece)
    if len(pcts) == 1:  # scheme-first / prose-separated, e.g. "3x2 back squat @ 70%"
        m = re.search(r"(\d+)\s*x\s*(\d+)", piece)
        pct = float(pcts[0]) / 100.0
        return [{"scheme": f"{m.group(1)}×{m.group(2)}" if m else "",
                 "pct": pct, "expectedLoad": round5(pct * mx)}]
    return [parse_note_segment(piece)]


def parse_sets_from_line(line, name, key, mx):
    # Parse the main lift's own sets, starting at its name (plus an immediately
    # preceding "NxM" scheme for scheme-first lines). Split on "·" up to the first
    # accessory / different-lift boundary; within a segment split "·"-joined sets and
    # "+"-joined sets (a "+" piece only counts as a set if it carries an "@ P%").
    m = re.search(re.escape(name) + r"\b", line, re.IGNORECASE)
    preceding = re.search(r"(\d+\s*x\s*\d+)\s*$", line[:m.start()])
    prefix = (preceding.group(1) + " ") if preceding else ""
    body = (prefix + line[m.end():].lstrip(":")).strip()
    sets, excluded, stopped = [], [], False
    for i, seg in enumerate(body.split("·")):
        seg = seg.strip()
        if not seg:
            continue
        if not stopped and i > 0:
            _, other_key = lift_at_start(seg)
            if STOP_RE.match(seg) or (other_key and other_key != key):
                stopped = True
        if stopped:
            excluded.append(seg)
            continue
        for j, piece in enumerate(re.split(r"\s\+\s", seg)):
            piece = piece.strip()
            if not piece:
                continue
            if j > 0 and not PCT_RE.search(piece):  # "+ accessory", not another set
                excluded.append(piece)
                continue
            sets.extend(parse_piece(piece, mx))
    return sets, " · ".join(excluded)


def lift_at_start(text):
    # (name, key) if text starts with a barbell lift name, else (None, None)
    stripped = text.strip()
    for name, key in LIFT_BY_NAME:
        if re.match(re.escape(name) + r"\b", stripped, re.IGNORECASE):
            return name, key
    return None, None


def identify_main_lift(session, load_cell):
    # Unify the two signals. If the headline (leadLift) differs from the load cell's
    # lift (backsolve), the cell belongs to an accessory -> main = leadLift (parsed
    # from its own line). Otherwise main = backsolve, parsed from its "@ P%" line.
    # Returns (section, anchor_line, excluded_text) or (None, None, None).
    back_key = backsolve_lift(load_cell)
    lead_line, lead_name, lead_key = find_lead_lift(session)
    if lead_key and lead_key != back_key:
        anchor, name, key = lead_line, lead_name, lead_key
        require_pct = False
    elif back_key:
        anchor, name = find_pct_line(session, back_key)
        if not anchor:
            return None, None, None
        key, require_pct = back_key, True
    else:
        return None, None, None
    mx = DEFAULT_MAXES[key]
    sets, excluded = parse_sets_from_line(anchor, name, key, mx)
    if require_pct and not any("pct" in st for st in sets):
        return None, None, None
    section = {"type": "mainLift", "label": "Main Lift",
               "liftName": DISPLAY_NAME[key], "liftKey": key, "sets": sets,
               "printedLoads": [[pct, load] for pct, load in parse_load_cell(load_cell)]}
    return section, anchor, excluded


def plan_year_for_month(month):
    # Plan runs Jul 2026 -> Jan 2027. Jan-Jun => 2027, Jul-Dec => 2026.
    return 2027 if month <= 6 else 2026


def parse_date_label(label):
    # label like "Mon Jul 20" ; returns (iso, dow)
    m = re.match(r"([A-Za-z]{3})\s+([A-Za-z]{3})\s+(\d+)", label.strip())
    dow, mon, day = m.group(1).upper(), m.group(2), int(m.group(3))
    year = plan_year_for_month(MONTHS[mon])
    iso = datetime.date(year, MONTHS[mon], day).isoformat()
    return iso, dow


def session_title(text):
    lines = text.strip().splitlines()
    if not lines:
        return ""
    first = re.split(r"\s—\s|\s·\s", lines[0].strip())[0].strip()
    return first.title() if first.isupper() else first


def extract_block_days():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    days = []
    for sheet in BLOCK_SHEETS:
        ws = wb[sheet]
        week_num, week_dates, week_focus = None, "", ""
        for row in ws.iter_rows(values_only=True):
            cells = [c for c in row]
            first = cells[0]
            if not any(c is not None for c in cells):
                continue
            if isinstance(first, str) and first.startswith("WEEK "):
                m = WEEK_HDR.match(first)
                if not m:
                    raise ValueError(f"Unparseable week header in {sheet!r}: {first!r}")
                week_num = int(m.group(1))
                week_dates = m.group(2).strip()
                week_focus = m.group(3).strip()
                continue
            if isinstance(first, str) and re.match(r"[A-Za-z]{3}\s+[A-Za-z]{3}\s+\d+", first):
                iso, dow = parse_date_label(first)
                session = (cells[2] or "").strip()
                load_cell = (cells[3] or "") if len(cells) > 3 and cells[3] else ""
                days.append({
                    "date": iso, "dow": dow, "dateLabel": first.strip(),
                    "block": BLOCK_LABEL[sheet], "week": week_num,
                    "weekDates": week_dates, "weekFocus": week_focus,
                    "sessionTitle": session_title(session),
                    "sections": [],
                    "_session": session,
                    "_loadCell": load_cell,
                })
    return days


def _prep_section(para):
    head = para.splitlines()[0]
    label, _, text = head.partition("—")
    return {
        "type": "prep", "label": label.strip() or "Prep",
        "text": (text.strip() + "\n" + "\n".join(para.splitlines()[1:])).strip(),
    }


def make_sections(session, main_lift, lead_line, excluded_text):
    paras = [p.strip() for p in re.split(r"\n\s*\n", session) if p.strip()]
    sections = []
    body_started = False
    for i, para in enumerate(paras):
        lines = para.splitlines()
        if i == 0:  # first line is the session title; drop it
            lines = lines[1:]
            if not lines:
                continue
        stripped = [ln.strip() for ln in lines]
        if main_lift and lead_line in stripped:
            idx = stripped.index(lead_line)
            pre = "\n".join(lines[:idx]).strip()
            if pre:
                sections.append(_prep_section(pre) if pre.lower().startswith("prep")
                                else {"type": "text",
                                      "label": "Accessory" if body_started else "Notes",
                                      "text": pre})
            sections.append(main_lift)
            body_started = True
            remainder = "\n".join(
                part for part in (excluded_text, "\n".join(lines[idx + 1:]).strip()) if part
            ).strip()
            if remainder:
                sections.append({"type": "text", "label": "Accessory", "text": remainder})
            continue
        if lines[0].lower().startswith("prep"):
            sections.append(_prep_section("\n".join(lines)))
            continue
        sections.append({
            "type": "text",
            "label": "Accessory" if body_started else "Notes",
            "text": "\n".join(lines),
        })
    if main_lift and main_lift not in sections:
        sections.append(main_lift)
    return sections


DOW_ORDER = ["MON", "TUE", "WED", "THU", "FRI", "SAT", "SUN"]
# Ramp-in runs Tue Jun 9 -> Sun Jul 19, 2026. Anchor to the REAL calendar: Jun 9
# 2026 is a TUESDAY (the block plan's Week 1 starts Mon Jul 20, a real Monday, so
# the ramp-in weekdays must line up with the real calendar too). R-weeks are the
# calendar weeks whose Mondays are Jun 8, 15, 22, 29, Jul 6, Jul 13.
RAMPIN_START = datetime.date(2026, 6, 9)
RAMPIN_END = datetime.date(2026, 7, 19)
RAMPIN_WEEK0_MONDAY = datetime.date(2026, 6, 8)


def extract_rampin_days():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["Ramp-In Jun9-Jul19"]
    rows = [[c for c in r] for r in ws.iter_rows(values_only=True)]
    one_thing = {}   # DOW -> (title, why, loads)
    week_dial = []   # ordered (Rn, dates, dial, checkpoint)
    for cells in rows:
        first = cells[0]
        if isinstance(first, str) and first.strip().upper() in DOW_ORDER:
            one_thing[first.strip().upper()] = (cells[1] or "", cells[2] or "", cells[3] or "")
        if isinstance(first, str) and re.match(r"R\d", first.strip()):
            week_dial.append((first.strip(), cells[1] or "", cells[2] or "", cells[3] or ""))
    days = []
    date = RAMPIN_START
    while date <= RAMPIN_END:
        dow = DOW_ORDER[date.weekday()]
        monday = date - datetime.timedelta(days=date.weekday())
        wi = (monday - RAMPIN_WEEK0_MONDAY).days // 7
        rn, wdates, dial, checkpoint = week_dial[wi]
        title, _why, loads = one_thing.get(dow, ("OFF", "", ""))
        one = title if not loads else f"{title}\n\nLoads / notes: {loads}"
        this_week = (dial + ("\nCheckpoint: " + checkpoint
                             if checkpoint and checkpoint != "—" else "")).strip()
        days.append({
            "date": date.isoformat(), "dow": dow,
            "dateLabel": date.strftime("%a %b %-d"),
            "block": "Ramp-In", "week": rn, "weekDates": wdates.strip(),
            "weekFocus": "Pre-plan garnish — one small thing before class",
            "sessionTitle": (title.split("—")[0].strip() if title else "Off"),
            "sections": [
                {"type": "text", "label": "The one thing (pre-class)", "text": one},
                {"type": "text", "label": f"This week · {rn}", "text": this_week},
            ],
        })
        date += datetime.timedelta(days=1)
    return days


def build_plan():
    days = extract_block_days()
    for d in days:
        main_lift, anchor, excluded = identify_main_lift(d["_session"], d["_loadCell"])
        d["sections"] = make_sections(d["_session"], main_lift, anchor, excluded)
    days = extract_rampin_days() + days
    days.sort(key=lambda d: d["date"])
    return {"lifts": LIFTS, "defaultMaxes": DEFAULT_MAXES, "days": days}


def strip_scratch(plan):
    for d in plan["days"]:
        for k in ("_session", "_loadCell"):
            d.pop(k, None)
    return plan


def main():
    plan = build_plan()
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(strip_scratch(plan), indent=2, ensure_ascii=False))
    print(f"Wrote {len(plan['days'])} days to {OUT}")


if __name__ == "__main__":
    main()
